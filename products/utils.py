import io
import csv
import openpyxl
from .models import Product, Inquiry, Category
from io import BytesIO

# Helper function to handle timezone-aware datetime fields


def convert_to_naive_datetime(dt):
    if dt and dt.tzinfo:  # Check if dt is not None and is timezone-aware
        return dt.astimezone(None)  # Convert to naive datetime
    return dt

# Function to format datetimes for Excel export


def format_datetime_for_excel(dt):
    if dt:
        # Format the datetime for Excel
        return dt.strftime('%Y-%m-%d %H:%M:%S')
    return ''  # Return an empty string if datetime is None

# Export Categories data to CSV


def export_categories_to_csv(queryset):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Name', 'Slug', 'Description', 'Created At'])

    for category in queryset:
        created_at = convert_to_naive_datetime(category.created_at)
        writer.writerow([
            category.name,
            category.slug,
            category.description,
            created_at if created_at is None else created_at.strftime(
                '%Y-%m-%d %H:%M:%S'),
        ])

    output.seek(0)
    return output.getvalue()

# Export Categories data to Excel


def export_categories_to_excel(queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Name', 'Slug', 'Description', 'Created At'])

    for category in queryset:
        created_at = convert_to_naive_datetime(category.created_at)
        ws.append([
            category.name,
            category.slug,
            category.description,
            format_datetime_for_excel(created_at),
        ])

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file

# Export Products data to CSV


def export_products_to_csv(queryset):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Name', 'Category', 'Stock',
                    'Available', 'Published', 'Created At'])

    for product in queryset:
        created_at = convert_to_naive_datetime(product.created_at)
        writer.writerow([
            product.name,
            product.category.name,
            product.stock,
            product.is_available,
            product.is_published,
            created_at if created_at is None else created_at.strftime(
                '%Y-%m-%d %H:%M:%S'),
        ])

    output.seek(0)
    return output.getvalue()

# Export Products data to Excel


def export_products_to_excel(queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Name', 'Category', 'Stock',
              'Available', 'Published', 'Created At'])

    for product in queryset:
        created_at = convert_to_naive_datetime(product.created_at)
        ws.append([
            product.name,
            product.category.name,
            product.stock,
            product.is_available,
            product.is_published,
            format_datetime_for_excel(created_at),
        ])

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file

# Export Inquiries data to CSV


def export_inquiries_to_csv(queryset):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Name', 'Email', 'Phone',
                    'Product', 'Category', 'Created At'])

    for inquiry in queryset:
        created_at = convert_to_naive_datetime(inquiry.created_at)
        writer.writerow([
            inquiry.name,
            inquiry.email,
            inquiry.phone,
            inquiry.product.name,
            inquiry.category.name,
            created_at if created_at is None else created_at.strftime(
                '%Y-%m-%d %H:%M:%S'),
        ])

    output.seek(0)
    return output.getvalue()

# Export Inquiries data to Excel


def export_inquiries_to_excel(queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Name', 'Email', 'Phone', 'Product', 'Category', 'Created At'])

    for inquiry in queryset:
        created_at = convert_to_naive_datetime(inquiry.created_at)
        ws.append([
            inquiry.name,
            inquiry.email,
            inquiry.phone,
            inquiry.product.name,
            inquiry.category.name,
            format_datetime_for_excel(created_at),
        ])

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file

# Export Reviews data to Excel


def export_reviews_to_excel(queryset):
    import pandas as pd
    from io import BytesIO

    data = []
    for review in queryset:
        data.append({
            'Product': review.product.name,
            'Name': review.name,
            'Email': review.email,
            'Rating': review.rating,
            'Review': review.review,
            'Created At': review.created_at,
        })
    
    df = pd.DataFrame(data)
    output = BytesIO()  # Use BytesIO to create an in-memory output
    df.to_excel(output, index=False)
    output.seek(0)  # Move to the beginning of the BytesIO buffer
    return output.getvalue()  # Return the Excel file content

# Export Reviews data to CSV


def export_reviews_to_csv(queryset):
    import csv
    from io import StringIO

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['Product', 'Name', 'Email', 'Rating', 'Review', 'Created At'])  # Header row

    for review in queryset:
        writer.writerow([review.product.name, review.name, review.email, review.rating, review.review, review.created_at])

    return output.getvalue()  # Return the CSV file content
