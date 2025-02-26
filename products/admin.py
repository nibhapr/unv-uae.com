from django.contrib import admin
from django.http import HttpResponse
from django.contrib.admin import DateFieldListFilter, SimpleListFilter
from .models import Category, Product, Inquiry, CategorySEO, ProductSEO, Review
from .utils import export_categories_to_csv, export_inquiries_to_csv, export_categories_to_excel, export_inquiries_to_excel, export_reviews_to_csv, export_reviews_to_excel
from import_export import resources, fields, widgets
from import_export.admin import ImportExportModelAdmin
from django import forms
from django.core.files import File
from openpyxl import load_workbook
import os
from import_export.widgets import ForeignKeyWidget, BooleanWidget
from django.utils import timezone
from datetime import datetime
from django.conf import settings
from django.utils.text import slugify


class CategorySEOInline(admin.StackedInline):
    model = CategorySEO
    can_delete = False
    verbose_name_plural = 'SEO Information'

    def get_max_num(self, request, obj=None, **kwargs):
        return 1


class ProductSEOInline(admin.StackedInline):
    model = ProductSEO
    can_delete = False
    verbose_name_plural = 'SEO Information'


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    inlines = [CategorySEOInline]
    list_display = ('name', 'slug', 'description',
                    'created_at', 'get_meta_title')
    prepopulated_fields = {'slug': ('name',)}
    search_fields = ('name', 'description')
    list_filter = [('created_at', DateFieldListFilter)]
    ordering = ('created_at',)
    actions = ['export_categories_to_csv_action',
               'export_categories_to_excel_action']

    def get_meta_title(self, obj):
        return obj.seo.meta_title if hasattr(obj, 'seo') else ''
    get_meta_title.short_description = 'Meta Title'

    def export_categories_to_csv_action(self, request, queryset):
        if 'apply_filter' in request.POST:
            pass
        csv_data = export_categories_to_csv(queryset)
        response = HttpResponse(csv_data, content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="categories.csv"'
        return response

    def export_categories_to_excel_action(self, request, queryset):
        excel_file = export_categories_to_excel(queryset)
        response = HttpResponse(
            excel_file,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="categories.xlsx"'
        return response

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        # Create or update SEO information
        CategorySEO.objects.get_or_create(category=obj)

    def get_inline_instances(self, request, obj=None):
        if not obj:
            return []
        return super().get_inline_instances(request, obj)


class ProductResource(resources.ModelResource):
    name = fields.Field(attribute='name', column_name='Name')
    category = fields.Field(
        attribute='category',
        column_name='Category',
        widget=ForeignKeyWidget(Category, 'name')
    )
    description = fields.Field(
        attribute='description', column_name='Description')
    stock = fields.Field(attribute='stock', column_name='Stock')
    is_available = fields.Field(
        attribute='is_available',
        column_name='Is Available',
        widget=BooleanWidget()
    )

    # Technical specifications
    max_resolution = fields.Field(
        attribute='max_resolution', column_name='Max Resolution')
    sensor = fields.Field(attribute='sensor', column_name='Sensor')
    day_night = fields.Field(attribute='day_night', column_name='Day Night')
    shutter = fields.Field(attribute='shutter', column_name='Shutter')
    adjustment_angle = fields.Field(
        attribute='adjustment_angle', column_name='Adjustment Angle')
    s_n = fields.Field(attribute='s_n', column_name='S/N')
    wdr = fields.Field(attribute='wdr', column_name='WDR')
    focal_length = fields.Field(
        attribute='focal_length', column_name='Focal Length')
    iris_type = fields.Field(attribute='iris_type', column_name='Iris Type')
    iris = fields.Field(attribute='iris', column_name='Iris')
    video_compression = fields.Field(
        attribute='video_compression', column_name='Video Compression')
    frame_rate = fields.Field(attribute='frame_rate', column_name='Frame Rate')
    video_bit_rate = fields.Field(
        attribute='video_bit_rate', column_name='Video Bit Rate')
    video_stream = fields.Field(
        attribute='video_stream', column_name='Video Stream')
    audio_compression = fields.Field(
        attribute='audio_compression', column_name='Audio Compression')
    two_way_audio = fields.Field(
        attribute='two_way_audio', column_name='Two Way Audio')
    suppression = fields.Field(
        attribute='suppression', column_name='Suppression')
    sampling_rate = fields.Field(
        attribute='sampling_rate', column_name='Sampling Rate')
    edge_storage = fields.Field(
        attribute='edge_storage', column_name='Edge Storage')
    network_storage = fields.Field(
        attribute='network_storage', column_name='Network Storage')
    protocols = fields.Field(attribute='protocols', column_name='Protocols')
    compatible_integration = fields.Field(
        attribute='compatible_integration', column_name='Compatible Integration')
    power = fields.Field(attribute='power', column_name='Power')
    dimensions = fields.Field(attribute='dimensions', column_name='Dimensions')
    weight = fields.Field(attribute='weight', column_name='Weight')
    material = fields.Field(attribute='material', column_name='Material')

    # Image fields and datetime fields remain the same
    photo_main = fields.Field(attribute='photo_main', column_name='Photo Main')
    photo_1 = fields.Field(attribute='photo_1', column_name='Photo 1')
    photo_2 = fields.Field(attribute='photo_2', column_name='Photo 2')
    photo_3 = fields.Field(attribute='photo_3', column_name='Photo 3')
    photo_4 = fields.Field(attribute='photo_4', column_name='Photo 4')

    list_date = fields.Field(
        column_name='List Date',
        attribute='list_date',
        widget=widgets.DateTimeWidget(format='%Y-%m-%d %H:%M:%S')
    )
    created_at = fields.Field(
        column_name='Created At',
        attribute='created_at',
        widget=widgets.DateTimeWidget(format='%Y-%m-%d %H:%M:%S')
    )

    class Meta:
        model = Product
        skip_unchanged = True
        report_skipped = True
        import_id_fields = ['name']
        fields = (
            'name', 'category', 'description', 'stock', 'is_available',
            'is_featured', 'is_published',
            'max_resolution', 'sensor', 'day_night', 'shutter',
            'adjustment_angle', 's_n', 'wdr', 'focal_length',
            'iris_type', 'iris', 'video_compression',
            'frame_rate', 'video_bit_rate', 'video_stream',
            'audio_compression', 'two_way_audio', 'suppression', 'sampling_rate',
            'edge_storage', 'network_storage',
            'protocols', 'compatible_integration',
            'power', 'dimensions', 'weight', 'material',
            'photo_main', 'photo_1', 'photo_2', 'photo_3', 'photo_4',
            'list_date', 'created_at'
        )


@admin.register(Product)
class ProductAdmin(ImportExportModelAdmin):
    resource_class = ProductResource
    inlines = [ProductSEOInline]
    list_display = ('name', 'category', 'stock', 'is_available', 'is_featured',
                    'is_published', 'created_at', 'seo_title', 'seo_rating')
    search_fields = ('name', 'description')
    actions = ['export_products_to_csv_action',
               'export_products_to_excel_action']
    change_list_template = 'admin/products/product/change_list.html'

    class AvailableFilter(SimpleListFilter):
        title = 'Availability'
        parameter_name = 'availability'

        def lookups(self, request, model_admin):
            return [
                ('True', 'Available'),
                ('False', 'Not Available'),
            ]

        def queryset(self, request, queryset):
            if self.value() == 'True':
                return queryset.filter(is_available=True)
            elif self.value() == 'False':
                return queryset.filter(is_available=False)
            return queryset

    list_filter = [
        ('category', admin.RelatedFieldListFilter),
        AvailableFilter,
        ('created_at', DateFieldListFilter),
        ('is_published', admin.BooleanFieldListFilter),
    ]

    def seo_title(self, obj):
        try:
            return obj.seo.title if hasattr(obj, 'seo') else ''
        except ProductSEO.DoesNotExist:
            return ''

    def seo_rating(self, obj):
        try:
            if hasattr(obj, 'seo'):
                return f"{obj.seo.rating} ({obj.seo.review_count} reviews)"
            return ''
        except ProductSEO.DoesNotExist:
            return ''

    def export_products_to_csv_action(self, request, queryset):
        if 'apply_filter' in request.POST:
            pass
        return self.export_to_csv(queryset, 'products.csv')

    def export_products_to_excel_action(self, request, queryset):
        return self.export_to_excel(queryset, 'products.xlsx')

    def export_to_csv(self, queryset, filename):
        import csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        writer.writerow(['ID', 'Name', 'Category', 'Stock',
                        'Is Available', 'Is Featured', 'Created At'])
        for product in queryset:
            writer.writerow([product.id, product.name, product.category.name, product.stock,
                            product.is_available, product.is_featured, product.created_at])
        return response

    def export_to_excel(self, queryset, filename):
        import pandas as pd
        data = []
        for product in queryset:
            data.append({
                'ID': product.id,
                'Name': product.name,
                'Category': product.category.name,
                'Stock': product.stock,
                'Is Available': product.is_available,
                'Is Featured': product.is_featured,
                'Created At': product.created_at,
            })
        df = pd.DataFrame(data)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        df.to_excel(response, index=False)
        return response

    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('import-products/', self.import_products_view,
                 name='import-products'),
        ]
        return custom_urls + urls

    def import_products_view(self, request):
        from django.shortcuts import render, redirect
        from django.contrib import messages
        import os

        if request.method == 'POST':
            form = ProductImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                image_base_path = form.cleaned_data['image_base_path']

                try:
                    wb = load_workbook(excel_file)
                    ws = wb.active

                    # Skip header row
                    rows = ws.iter_rows(min_row=2)

                    for row in rows:
                        try:
                            name = row[0].value
                            category_name = row[1].value
                            description = row[2].value
                            stock = row[3].value
                            image_path = row[4].value if len(row) > 4 else None

                            # Ensure category_name is not None or empty
                            if not category_name:
                                raise ValueError(
                                    f"Category name is required for product: {name}")

                            # Get or create category with error handling
                            try:
                                category, created = Category.objects.get_or_create(
                                    name=category_name.strip()
                                )
                            except Exception as e:
                                raise ValueError(
                                    f"Failed to create category '{category_name}': {str(e)}")

                            # Create product with explicit category assignment
                            product = Product(
                                name=name,
                                category=category,  # Explicitly assign category
                                description=description,
                                stock=stock or 0  # Default to 0 if None
                            )
                            product.save()

                            # Handle image upload
                            if image_path:
                                full_path = os.path.join(
                                    image_base_path, image_path)
                                if os.path.exists(full_path):
                                    with open(full_path, 'rb') as f:
                                        product.photo_main.save(
                                            os.path.basename(image_path), File(f))

                        except Exception as e:
                            messages.error(
                                request, f"Error importing row {row[0].row}: {str(e)}")
                            continue

                    messages.success(request, 'Products imported successfully')
                    return redirect('..')

                except Exception as e:
                    messages.error(request, f'Error importing file: {str(e)}')
            else:
                messages.error(request, 'Invalid form submission')

        return render(request, 'admin/import_products.html', {'form': form})

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        # Create or update SEO information
        ProductSEO.objects.get_or_create(product=obj)


@admin.register(Inquiry)
class InquiryAdmin(admin.ModelAdmin):
    list_display = ('name', 'email', 'phone', 'message',
                    'category', 'product', 'created_at')
    search_fields = ('name', 'email', 'message')
    actions = ['export_inquiries_to_csv_action',
               'export_inquiries_to_excel_action']

    class EmailListFilter(admin.SimpleListFilter):
        title = 'Email'
        parameter_name = 'email'

        def lookups(self, request, model_admin):
            return [
                ('@gmail.com', 'Gmail'),
                ('@yahoo.com', 'Yahoo'),
                ('@hotmail.com', 'Hotmail'),
            ]

        def queryset(self, request, queryset):
            if self.value():
                return queryset.filter(email__icontains=self.value())
            return queryset

    list_filter = [
        ('created_at', admin.DateFieldListFilter),
        EmailListFilter,
    ]

    def save_inquiry(self, request, form, change):
        inquiry = form.save(commit=False)
        inquiry.save()

    def export_inquiries_to_csv_action(self, request, queryset):
        if 'select_all' in request.POST:
            queryset = Inquiry.objects.all()
        csv_data = export_inquiries_to_csv(queryset)
        response = HttpResponse(csv_data, content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="inquiries.csv"'

        for inquiry in queryset:
            self.save_inquiry(request, inquiry)

        return response

    def export_inquiries_to_excel_action(self, request, queryset):
        if 'select_all' in request.POST:
            queryset = Inquiry.objects.all()
        excel_file = export_inquiries_to_excel(queryset)
        response = HttpResponse(
            excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="inquiries.xlsx"'

        for inquiry in queryset:
            self.save_inquiry(request, inquiry)

        return response

    export_inquiries_to_csv_action.short_description = "Export selected inquiries to CSV"
    export_inquiries_to_excel_action.short_description = "Export selected inquiries to Excel"


@admin.register(Review)
class ReviewAdmin(admin.ModelAdmin):
    list_display = ('product', 'name', 'email', 'rating', 'created_at')
    list_filter = ('rating', 'created_at')
    search_fields = ('name', 'email', 'review', 'product__name')
    readonly_fields = ('created_at',)
    date_hierarchy = 'created_at'
    actions = ['export_reviews_to_csv_action',
               'export_reviews_to_excel_action']

    fieldsets = (
        ('Review Information', {
            'fields': ('product', 'rating', 'review')
        }),
        ('Reviewer Details', {
            'fields': ('name', 'email', 'phone')
        }),
        ('Metadata', {
            'fields': ('created_at',),
            'classes': ('collapse',)
        }),
    )

    def get_readonly_fields(self, request, obj=None):
        if obj:
            return self.readonly_fields + ('created_at',)
        return self.readonly_fields

    def export_reviews_to_csv_action(self, request, queryset):
        if 'select_all' in request.POST:
            queryset = Review.objects.all()
        csv_data = export_reviews_to_csv(queryset)
        response = HttpResponse(csv_data, content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="reviews.csv"'
        return response

    def export_reviews_to_excel_action(self, request, queryset):
        if 'select_all' in request.POST:
            queryset = Review.objects.all()
        excel_file = export_reviews_to_excel(queryset)
        response = HttpResponse(
            excel_file,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="reviews.xlsx"'
        return response


class ProductImportForm(forms.Form):
    excel_file = forms.FileField(label='Excel File')
    image_base_path = forms.CharField(
        label='Base Path for Images',
        help_text='Enter the base directory path where your images are stored',
        required=True
    )
